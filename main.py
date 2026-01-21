from fastapi import FastAPI, File, UploadFile, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from jinja2 import Template
from datetime import datetime

app = FastAPI()

@app.post("/excel")
def crear_excel(payload: dict):
    data = payload.get("data", [])
    if not data:
        return {"error": "No data provided"}

    filename = payload.get("filename", "archivo.xlsx")
    sheet = payload.get("sheet", "Datos")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if filename.lower().endswith(".xlsx"):
        base = filename[:-5]
    else:
        base = filename
    
    final_filename = f"{base}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    headers = list(data[0].keys())

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    # 🔹 NUEVO: estilo para primera columna
    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )
    bold_font = Font(bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header))

    # 🔹 NUEVO: aplicar estilo a A2:A...
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.fill = green_fill
        cell.font = bold_font

    output = BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    excel_bytes = output.getvalue()

    return StreamingResponse(
        BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{final_filename}"',
            "Content-Length": str(len(excel_bytes)),
            "Cache-Control": "no-store"
        }
    )

html_template = """
<p style="font-weight: bold; margin-bottom: 10px;">
  {{ archivo }}
</p>
<table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse;">
  <thead>
    <tr>
      {% for header in headers %}
      <th style="background-color: {{ header.bg_color }}; color: {{ header.font_color }}; font-weight: {{ header.font_weight }};">
        {{ header.value }}
      </th>
      {% endfor %}
    </tr>
  </thead>
  <tbody>
    {% for row in rows %}
    <tr>
      {% for cell in row %}
      <td style="background-color: {{ cell.bg_color }}; color: {{ cell.font_color }}; font-weight: {{ cell.font_weight }};">
        {{ cell.value }}
      </td>
      {% endfor %}
    </tr>
    {% endfor %}
  </tbody>
</table>
{% if info_text %}
<p style="font-weight: bold; margin-bottom: 10px;">
  {{ info_text }}
</p>
{% endif %}
"""

def get_cell_styles(cell):
    bg_color = "transparent"

    if cell.fill and cell.fill.fill_type == "solid":
        color = cell.fill.start_color
        if color and color.rgb:
            bg_color = f"#{str(color.rgb)[-6:]}"

    font_color = "black"
    if cell.font and cell.font.color:
        color = cell.font.color
        if color and color.rgb:
            font_color = f"#{str(color.rgb)[-6:]}"

    font_weight = "bold" if cell.font and cell.font.bold else "normal"

    return {
        "bg_color": bg_color,
        "font_color": font_color,
        "font_weight": font_weight,
    }

COLUMNAS_PERMITIDAS = {"Nombre", "Correo", "Telefono"}

@app.post("/procesar_excel", response_class=HTMLResponse)
async def procesar_excel(file: UploadFile = File(...), limit: int = Query(6, ge=1, le=10)):
    contents = await file.read()
    wb = load_workbook(BytesIO(contents), data_only=True)
    ws = wb.active

    # Total de registros (sin header)
    total_registros = ws.max_row - 1
    total_registros = max(total_registros, 0)
    
    # Leer headers y mapear índice
    header_cells = list(ws[1])
    headers = []
    columnas_idx = []

    for idx, cell in enumerate(header_cells):
        if cell.value in COLUMNAS_PERMITIDAS:
            columnas_idx.append(idx)
            headers.append({
                "value": cell.value or "",
                **get_cell_styles(cell)
            })

    # Filas limitadas dinámicamente
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        if i > limit:
            break

        fila = []
        for idx in columnas_idx:
            cell = row[idx]
            fila.append({
                "value": cell.value or "",
                **get_cell_styles(cell)
            })
        rows.append(fila)
    
    # Texto informativo
    info_text = f"Mostrando {len(rows)} de {total_registros} registros"
    html = Template(html_template).render(headers=headers, rows=rows, info_text=info_text, archivo=file.filename)
    return html


















